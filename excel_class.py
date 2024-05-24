from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from settings.config import config_json
from time import sleep
from utilities import chi2_test


class ExcelFile:
    def __init__(self, file_name: str, tables_type: int):
        self.file_type = tables_type
        if self.file_type not in range(0,5):
            raise ValueError("Wrong table type. tables_type is in (0,1,2,3,4)")
        self.file_name = file_name
        self.row = 1
        self.space = 2
        self.config_style = dict(config_json)
        self.workbook = Workbook()
        self.worksheet = None
        if tables_type != 0:
            self.chi_square_test = False
        else:
            self.chi_square_test = True

    def add_sheet(self, sheet_name: str, data: dict, main_var_name: str, footer_text=None) -> None:
        self.__normalize_sheet_name(sheet_name)
        secondary_var_name = list(data.keys())[0]
        df = data[secondary_var_name]
        len_df = len(df.columns)

        if self.sheet_name in self.workbook.sheetnames:
            self.worksheet = self.workbook[self.sheet_name]
        else:
            self.worksheet = self.workbook.create_sheet(title=self.sheet_name)

        # Header
        header_cell = self.worksheet.cell(row=self.row, column=1)
        header_cell.value = main_var_name
        self.worksheet.merge_cells(start_row=self.row, end_row=self.row, start_column=1, end_column=len_df)
        header_cell.alignment = Alignment(horizontal=self.config_style["Header"]["Alignment"]["horizontal"],
                                          vertical=self.config_style["Header"]["Alignment"]["vertical"])
        header_cell.font = Font(name=self.config_style["Header"]["Font"]["name"],
                                size=self.config_style["Header"]["Font"]["size"],
                                bold=self.config_style["Header"]["Font"]["bold"],
                                color=self.config_style["Header"]["Font"]["color"])

        for column in range(1, len_df + 1):
            cell = self.worksheet.cell(row=self.row, column=column)
            cell.border = Border(top=Side(border_style=self.config_style["Header"]["Border"]["border_style"],
                                          color=self.config_style["Header"]["Border"]["color"]),
                                 bottom=Side(border_style=self.config_style["Header"]["Border"]["border_style"],
                                             color=self.config_style["Header"]["Border"]["color"]))

        # Sub header
        if self.file_type != 4:
            sub_header_cell = self.worksheet.cell(row=self.row + 1, column=3)
            sub_header_cell.value = secondary_var_name
            self.worksheet.merge_cells(start_row=self.row + 1, end_row=self.row + 1, start_column=3, end_column=len_df)
            sub_header_cell.alignment = Alignment(horizontal=self.config_style["Sub_Header"]["Alignment"]["horizontal"],
                                                  vertical=self.config_style["Sub_Header"]["Alignment"]["vertical"])
            sub_header_cell.font = Font(name=self.config_style["Sub_Header"]["Font"]["name"],
                                        size=self.config_style["Sub_Header"]["Font"]["size"],
                                        bold=self.config_style["Sub_Header"]["Font"]["bold"],
                                        color=self.config_style["Sub_Header"]["Font"]["color"])
        else:
            self.row -= 1

        # Body
        self.worksheet.append(df.columns.tolist())
        for _, row in df.iterrows():
            self.worksheet.append(row.tolist())

        # Vars Header
        for row in self.worksheet.iter_rows(min_row=self.row + 2, max_row=self.row + 2, min_col=1, max_col=len_df):
            for cell in row:
                cell.fill = PatternFill(start_color=self.config_style["Var_Header"]["PatternFill"]["start_color"],
                                        end_color=self.config_style["Var_Header"]["PatternFill"]["end_color"],
                                        fill_type=self.config_style["Var_Header"]["PatternFill"]["fill_type"])
                cell.font = Font(bold=self.config_style["Var_Header"]["Font"]["bold"],
                                 name=self.config_style["Var_Header"]["Font"]["name"],
                                 size=self.config_style["Var_Header"]["Font"]["size"],
                                 color=self.config_style["Var_Header"]["Font"]["color"])
                cell.alignment = Alignment(horizontal=self.config_style["Var_Header"]["Alignment"]["horizontal"],
                                           vertical=self.config_style["Var_Header"]["Alignment"]["vertical"])
                cell.border = Border(top=Side(border_style=self.config_style["Var_Header"]["Border"]["border_style"],
                                              color=self.config_style["Var_Header"]["Border"]["color"]),
                                     bottom=Side(border_style=self.config_style["Var_Header"]["Border"]["border_style"],
                                                 color=self.config_style["Var_Header"]["Border"]["color"]))

        # Vars
        for row in self.worksheet.iter_rows(min_row=self.row + 3, max_row=self.worksheet.max_row, min_col=1,
                                            max_col=len_df):
            for cell in row:
                cell.font = Font(color=self.config_style["Body"]["Font"]["color"],
                                 name=self.config_style["Body"]["Font"]["name"],
                                 size=self.config_style["Body"]["Font"]["size"])
                cell.alignment = Alignment(horizontal=self.config_style["Body"]["Alignment"]["horizontal"],
                                           vertical=self.config_style["Body"]["Alignment"]["vertical"])
                if cell.column != 1:
                    if self.file_type == 4:
                        if cell.column == len_df:
                            cell.number_format = "0.0%"
                        else:
                            cell.number_format = "#,##0"

                    elif self.file_type != 0:
                        cell.number_format = '0.0%'

                    else:
                        cell.number_format = "#,##0"
                if cell.row == self.worksheet.max_row and self.file_type != 3:
                    for column in range(1, len_df + 1):
                        cell = self.worksheet.cell(row=self.worksheet.max_row, column=column) 
                        cell.border = Border(top=Side(border_style=self.config_style["Footer"]["Border"]["border_style"],
                                               color=self.config_style["Footer"]["Border"]["color"]))
                        cell.font = Font(bold=True)
                        
        # Footer
        if footer_text:
            footer_cell = self.worksheet.cell(row=self.worksheet.max_row + 1, column=1)
            footer_cell.value = footer_text
            self.worksheet.merge_cells(start_row=self.worksheet.max_row,
                                       end_row=self.worksheet.max_row,
                                       start_column=1, end_column=len_df)
            footer_cell.alignment = Alignment(horizontal=self.config_style["Footer"]["Alignment"]["horizontal"],
                                              vertical=self.config_style["Footer"]["Alignment"]["vertical"])
            footer_cell.font = Font(color=self.config_style["Footer"]["Font"]["color"],
                                    name=self.config_style["Footer"]["Font"]["name"],
                                    size=self.config_style["Footer"]["Font"]["size"])
            self.row += 1
            for column in range(1, len_df + 1):
                cell = self.worksheet.cell(row=self.worksheet.max_row, column=column)
                cell.border = Border(top=Side(border_style=self.config_style["Footer"]["Border"]["border_style"],
                                              color=self.config_style["Footer"]["Border"]["color"]))

        # Chi-Square test
        if self.chi_square_test:
            result = chi2_test(df)
            if result[0]:
                header_cell.value = main_var_name + "*"
            chi_square_cell = self.worksheet.cell(row=self.worksheet.max_row + 1, column=1)
            chi_square_cell.value = str(result[1])
            self.worksheet.merge_cells(start_row=self.worksheet.max_row,
                                       end_row=self.worksheet.max_row,
                                       start_column=1, end_column=len_df)
            chi_square_cell.alignment = Alignment(horizontal=self.config_style["Footer"]["Alignment"]["horizontal"],
                                                  vertical=self.config_style["Footer"]["Alignment"]["vertical"])
            chi_square_cell.font = Font(color=self.config_style["Footer"]["Font"]["color"],
                                        name=self.config_style["Footer"]["Font"]["name"],
                                        size=self.config_style["Footer"]["Font"]["size"])

            p_val_cell = self.worksheet.cell(row=self.worksheet.max_row + 1, column=1)
            p_val_cell.value = str(result[2])
            self.worksheet.merge_cells(start_row=self.worksheet.max_row,
                                       end_row=self.worksheet.max_row,
                                       start_column=1, end_column=len_df)
            p_val_cell.alignment = Alignment(horizontal=self.config_style["Footer"]["Alignment"]["horizontal"],
                                             vertical=self.config_style["Footer"]["Alignment"]["vertical"])
            p_val_cell.font = Font(color=self.config_style["Footer"]["Font"]["color"],
                                   name=self.config_style["Footer"]["Font"]["name"],
                                   size=self.config_style["Footer"]["Font"]["size"])

            self.row += self.space + 5 + len(df)
        else:
            self.row += self.space + 3 + len(df)

    def __normalize_sheet_name(self, sheet_name: str) -> None:
        drop_character = ("\\", "/", "?", "*", "[", "]")
        for character in drop_character:
            sheet_name.replace(character, '')
        if len(sheet_name) > 25:
            self.sheet_name = sheet_name[0:25]
        else:
            self.sheet_name = sheet_name

    def __resize_columns(self):
        for column in self.worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length

            adjusted_width = max_length + 2
            self.worksheet.column_dimensions[column_letter].width = adjusted_width

    def restart(self) -> None:
        self.__resize_columns()
        self.row = 1

    def save(self) -> None:
        if "Sheet" in self.workbook.sheetnames:
            self.workbook.remove(self.workbook["Sheet"])

        count = 0
        while count < 5:
            try:
                self.workbook.save("results//" + self.file_name)
                break
            except PermissionError:
                print(f"Permissions denied. Please close the file: results//{self.file_name}")
                count += 1
                sleep(5)
        else:
            print("Permissions denied, execution interrupted")
